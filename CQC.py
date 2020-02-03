from openpyxl import *
from openpyxl import *

def get_data_order(filename,dealtype):
	wb = load_workbook(filename)
	ws1 = wb['Sheet1']
	if 'Order' in wb.sheetnames:
		wb.remove(wb['Order'])
		ws2 = wb.create_sheet('Order')
	else:
		ws2 = wb.create_sheet('Order')
	work_row = int(ws1.max_row/5)+1 
	work_column = int(ws1.max_column)+1
	for j in range(1,work_column):
			ws2.cell(row=1,column=j).value = str(ws1.cell(row=1,column=j).value).strip()
	for i in range(1,work_row):
		for j in range(1,work_column):
			ws2.cell(row=i+1,column=j).value = str(ws1.cell(row=((i-1)*5+2),column=j).value).strip()	
	for i in range(1,work_row):
		if dealtype.lower() == 'ccc':
			factory = str(ws2.cell(row=i+1,column=2).value)
			ws2.cell(row=i+1,column=work_column).value = factory[6:10]
		else:
			ws2.cell(row=i+1,column=work_column).value = ws2.cell(row=i+1,column=3).value
	wb.save(filename)

def get_data_fine(filename,ftpye):
	wb = load_workbook(filename)
	ws1 = wb['Order']
	if ftype == 1:
		ttype == '有效'
	else：
		ttype == '暂停'
	if '一致性' in wb.sheetnames:
		wb.remove(wb['一致性'])
		ws2 = wb.create_sheet('一致性')
	else:
		ws2 = wb.create_sheet('一致性')
	work_row = int(ws1.max_row) + 1
	worklist = list(range(2,work_row)) #序号
	YZX = 1
	total = 2
	ws2.cell(row=1,column=1).value = '一致性检查数量'
	ws2.cell(row=1,column=2).value = '序号'
	ws2.cell(row=1,column=3).value = '证书编号'
	ws2.cell(row=1,column=4).value = '持证人'
	ws2.cell(row=1,column=5).value = '制造商'
	ws2.cell(row=1,column=6).value = '产品类别'
	while (worklist != []) and (total <= work_row):
		worklist1 = []
		YZX_row = worklist[0]
		worklist.remove(YZX_row)
		if ws1.cell(row=YZX_row,column=15).value == ttype:
			ws2.cell(row=total,column=1).value = YZX
			ws2.cell(row=total,column=2).value = ws1.cell(row=YZX_row,column=1).value
			ws2.cell(row=total,column=3).value = ws1.cell(row=YZX_row,column=2).value
			ws2.cell(row=total,column=4).value = ws1.cell(row=YZX_row,column=4).value
			ws2.cell(row=total,column=5).value = ws1.cell(row=YZX_row,column=5).value
			ws2.cell(row=total,column=6).value = ws1.cell(row=YZX_row,column=3).value
			ws2.cell(row=total,column=7).value = ws1.cell(row=YZX_row,column=13).value
			total += 1
			#same manufacturer
			for therow in worklist:
				if (ws1.cell(row=therow,column=5).value == ws1.cell(row=YZX_row,column=5).value) and (ws1.cell(row=therow,column=3).value == ws1.cell(row=YZX_row,column=3).value):
					if ws1.cell(row=therow,column=15).value == ttype:
						worklist1.append(therow)							
			while worklist1 != []:
				for therow in worklist1:			
					ws2.cell(row=total,column=2).value = ws1.cell(row=therow,column=1).value
					ws2.cell(row=total,column=3).value = ws1.cell(row=therow,column=2).value
					ws2.cell(row=total,column=4).value = ws1.cell(row=therow,column=4).value
					ws2.cell(row=total,column=5).value = ws1.cell(row=therow,column=5).value
					ws2.cell(row=total,column=6).value = ws1.cell(row=therow,column=3).value
					ws2.cell(row=total,column=7).value = ws1.cell(row=therow,column=13).value
					worklist.remove(therow)
					total += 1
				worklist1 = []		
			YZX += 1
	ws2.cell(row=work_row+2,column=1).value = '一致性个数：'
	ws2.cell(row=work_row+2,column=2).value = YZX-1
	wb.save(filename)

def get_FEE(filename,dealtype):		
	wb = load_workbook(filename)
	ws1 = wb['Order']
	ws2 = wb['一致性']	
	work_row = int(ws1.max_row)+1
	FEE = 0	
	if dealtype.lower() == 'ccc':
		for i in range(1,work_row):
			IssueDate = str(ws1.cell(row=i,column=11).value)
			if IssueDate[0:4] != '2019':
				FEE += 100 
	#else:
	ws2.cell(row=work_row+3,column=1).value = '收费证书费'
	ws2.cell(row=work_row+3,column=2).value = FEE		
	wb.save(filename)
	
def get_ODMOEM(filename):
	wb = load_workbook(filename)
	ws1 = wb['Order']
	ws2 = wb['一致性']
	work_row = int(ws1.max_row)+1
	ODM = []
	OEM = []
	odmoemvalue = []
	i = work_row+4
	ws2.cell(row=i,column=1).value = 'ODM证书：'
	ws2.cell(row=i,column=2).value = 'ODM证书派生数：'
	ws2.cell(row=i,column=3).value = 'ODM证书派证书：'
	for j in range(1,work_row):
		if ws1.cell(row=j,column=13).value == 'ODM' or ws1.cell(row=j,column=13).value == '已获证书结果模式':
			ODM.append(str((ws1.cell(row=j,column=14)).value))
	odmoemvalue = set(ODM)
	i += 1
	for odmoem in odmoemvalue:
			if odmoem != 'None':
				ws2.cell(row=i,column=1).value = odmoem
				ws2.cell(row=i,column=2).value = ODM.count(odmoem)
				x = 3
				for j in range(1,work_row):
					if ws1.cell(row=j,column=14).value == odmoem:
						ws2.cell(row=i,column=x).value = ws1.cell(row=j,column=2).value
						x += 1
				i += 1			
	wb.save(filename)		


filename = input('Input file-name to process:')
dealtype = input('CCC or CQC?:')
ftpye = input('1 for 监督，2 for 恢复:')
get_data_order(filename,dealtype)
get_data_fine(filename, ftpye)
get_FEE(filename,dealtype)
get_ODMOEM(filename)